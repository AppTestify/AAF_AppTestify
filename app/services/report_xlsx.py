"""Excel/XLSX report builders."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def _autosize_sheet(ws) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[letter].width = min(max_len + 2, 48)


def _write_table_sheet(ws, rows: list[dict[str, Any]], *, title: str | None = None) -> None:
    if title:
        ws.append([title])
        ws["A1"].font = Font(bold=True)
        ws.append([])
    if not rows:
        ws.append(["No data"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def build_runs_summary_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Governance run summary"])
    summary["A1"].font = Font(bold=True)
    summary.append(["Total runs", len(rows)])
    succeeded = sum(1 for r in rows if r.get("status") == "succeeded")
    failed = sum(1 for r in rows if r.get("status") == "failed")
    summary.append(["Succeeded", succeeded])
    summary.append(["Failed", failed])

    data = wb.create_sheet("Runs")
    _write_table_sheet(data, rows)
    _autosize_sheet(data)
    _autosize_sheet(summary)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_audit_events_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "AuditEvents"
    _write_table_sheet(ws, rows, title="Audit events export")
    _autosize_sheet(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_portfolio_executive_xlsx(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Executive portfolio report"])
    summary["A1"].font = Font(bold=True)
    for key in (
        "projects_total",
        "active_projects",
        "releases_total",
        "releases_approved",
        "releases_blocked",
        "high_risk_open",
        "avg_confidence",
        "avg_consensus",
    ):
        summary.append([key, report.get(key)])

    projects = wb.create_sheet("Projects")
    breakdown = report.get("project_breakdown") or []
    if breakdown:
        headers = list(breakdown[0].keys())
        projects.append(headers)
        for row in breakdown:
            projects.append([row.get(h) for h in headers])
    else:
        projects.append(["No project breakdown"])
    _autosize_sheet(projects)
    _autosize_sheet(summary)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
